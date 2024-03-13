import os

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

filename = input("Адрес файла: ")

try:
    workbook = load_workbook(filename)

    if not os.path.exists(filename):
        raise FileNotFoundError(f"Файл {filename} не найден.")

    sheet = workbook['Лист1']

    start_column = column_index_from_string('DJ')
    end_column = column_index_from_string('HT')

    for col in range(end_column, start_column - 1, -1):
        sheet.delete_cols(col)

    row = 2

    while sheet[f'A{row}'].value:

        column = 29

        is_graph = sheet.cell(row=1, column=column).value

        while is_graph.startswith('Платеж'):

            cell = sheet.cell(row=row, column=column).value

            if cell == 0:

                sheet.cell(row=row, column=column).value = "0,00;"

            else:

                parts = tuple(cell.split(';', 1))
                data_with_dot = parts[0].replace(',', '.')
                number = round(float(data_with_dot), 2)

                sheet.cell(row=row, column=column).value = str(number) + ";" + parts[1]

            column += 1
            is_graph = sheet.cell(row=1, column=column).value

        row += 1

    filename_corr = filename[:-4] + "_corr." + filename[-4:]

    workbook.save(filename_corr)

    print(f"Готово. Исправленный файл сохранен как {filename_corr}")

except FileNotFoundError as e:
    print(e)
except Exception as e:
    print(f"Произошла ошибка: {e}")
